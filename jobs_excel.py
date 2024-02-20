"""Module for data processing from Excel files used for this project."""


from openpyxl import load_workbook


def get_jobs(filename):
    """Gets job data from an Excel file. Order of data is ordered to match order of database columns.

    Keyword arguments:
    filename -- Excel xlsx file storing job data

    Returns:
    jobs -- Job data from Excel file
    """

    workbook = load_workbook(filename=filename)
    sheet = workbook.active

    jobs = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Shaping tuple by how the database takes it
        # Database:
        # (job_id, title, company_name, location, description, posted_at, salary, remote, links, qualifications)
        # Excel file:
        # Company Name, Posting Age, Job ID, Country, Location, Publication Date, Salary Max, Salary Min, Salary Type,
        #   Job Title

        # Make salary string
        if row[7] == row[6]:
            salary = str(row[7])
        else:
            salary = f"{row[7]} - {row[6]}"
        if row[8] != "N/A":
            salary += f" {row[8]}"

        jobs += [(row[2], row[9], row[0], row[4], "", row[1], salary, "", [], [])]

    return jobs
