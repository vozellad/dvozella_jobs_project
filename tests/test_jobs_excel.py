"""Testing functions in test_jobs_results.py."""

import pytest
import jobs_excel
from openpyxl import load_workbook


def test_get_jobs():
    # Requires Excel file to be in tests directory
    workbook = load_workbook(filename="test_Sprint3Data.xlsx")
    sheet = workbook.active

    rows = [row for row in sheet.iter_rows(min_row=2, values_only=True)]
    assert len(rows) <= 750
